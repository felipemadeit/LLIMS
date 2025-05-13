from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell.cell import MergedCell
import re  # Agrega la importación de re


def merged_cell(ws, celda_coord, valor) -> bool:
    """
    Escribe un valor en una celda específica en una hoja de Excel, manejando celdas combinadas.

    Esta función escribe un valor en una coordenada de celda específica en una hoja de Excel.
    Cuando la celda objetivo es parte de un rango combinado, la función identifica la celda principal
    (celda superior izquierda) del rango combinado y escribe el valor allí, manteniendo el comportamiento
    correcto de Excel.

    Args:
        ws: El objeto hoja de trabajo (openpyxl Worksheet)
        celda_coord (str): Coordenada de celda en formato Excel (ej. 'A1', 'B12')
        valor: El valor a escribir en la celda

    Returns:
        bool: True si la operación de escritura fue exitosa, False en caso contrario
    """
    try:
        # Analizar la coordenada de la celda usando regex para extraer letras de columna y número de fila
        match = re.match(r'([A-Za-z]+)(\d+)', celda_coord)
        if not match:
            print(f"Formato de coordenada incorrecto: {celda_coord}")
            return False

        # Extraer letras de columna y número de fila del match
        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)

        # Verificar si la celda está en un rango combinado
        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col

            # Verificar si la celda objetivo está dentro de este rango combinado
            if min_row <= row <= max_row and min_col <= col <= max_col:
                # Obtener la celda principal (superior izquierda) del rango combinado
                celda_principal = ws.cell(row=min_row, column=min_col)
                # Escribir el valor en la celda principal
                celda_principal.value = valor
                print(
                    f"Escribiendo '{valor}' en la celda principal {get_column_letter(min_col)}{min_row} del rango combinado")
                return True

        # Si no está en un rango combinado, escribir directamente
        # NOTA: No es necesario verificar si es una MergedCell aquí,
        # ya que ya hemos verificado todos los rangos combinados
        ws.cell(row=row, column=col).value = valor
        print(f"Escribiendo '{valor}' directamente en {celda_coord}")
        return True

    except Exception as e:
        # Capturar cualquier error inesperado que pueda ocurrir
        print(f"Error en write_cell con coordenada {celda_coord}: {str(e)}")
        return False


def write_header_data(wb_destiny, header_data):
    """
    Escribe los datos del encabezado en la hoja 'Reporte' del libro de trabajo.

    Args:
        wb_destiny: El libro de trabajo de destino
        header_data: Lista con los datos del encabezado
    """
    sheet_to_write = wb_destiny["Reporte"]

    try:
        # Extraer datos del encabezado
        company_name = header_data[0]
        client_name = header_data[0]  # Parece ser el mismo que company_name
        client_address = header_data[1]
        lab_received_date = header_data[9]
        project_location = header_data[8]  # No usado actualmente
        client_phone = header_data[4]

        # Escribir datos en celdas (posiblemente combinadas)
        merged_cell(sheet_to_write, "K7", company_name)
        merged_cell(sheet_to_write, "K8", client_name)
        merged_cell(sheet_to_write, "K9", client_address)
        merged_cell(sheet_to_write, "AK6", lab_received_date)
        merged_cell(sheet_to_write, "AK8", client_name)
        merged_cell(sheet_to_write, "AK9", client_phone)

        print("Datos del encabezado escritos correctamente")
        return True

    except Exception as ex:
        print(f"ERROR en write_header_data: {ex}")
        return False


# Ejemplo de uso
def main():
    try:
        # Cargar un libro de trabajo
        wb = load_workbook('tu_archivo.xlsx')

        # Preparar algunos datos de ejemplo para el encabezado
        # (ajustar según la estructura real de tus datos)
        header_data = [
            "Empresa ABC",  # 0: company_name/client_name
            "Calle Principal 123",  # 1: client_address
            "Ciudad",  # 2
            "Estado",  # 3
            "123-456-7890",  # 4: client_phone
            "correo@ejemplo.com",  # 5
            "Código postal",  # 6
            "País",  # 7
            "Ubicación del proyecto",  # 8: project_location
            "2023-05-12"  # 9: lab_received_date
        ]

        # Escribir los datos del encabezado
        write_header_data(wb, header_data)

        # Guardar el libro de trabajo
        wb.save('tu_archivo_actualizado.xlsx')
        print("Archivo guardado correctamente")

    except Exception as e:
        print(f"ERROR en main: {e}")

