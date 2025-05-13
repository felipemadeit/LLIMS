from openpyxl.reader.excel import load_workbook

from Format.footer_format_copy import footer_format_copy
from Format.header_analitic_format_copy import header_analitic_format_copy
from Format.header_format_copy import header_format_copy
from Format.lab_format_copy import lab_format_copy
from Read.excel_chain_data_reader import excel_chain_data_reader
from Read.excel_header_reader import excel_header_reader
from Read.matrix_data_reader import matrix_data_reader
from Utils.apply_font_to_worksheet import apply_font_to_worksheet
from Utils.set_height_for_all_rows import set_height_for_all_rows
from Write.write_header_data import write_header_data
from Write.write_lab_data import write_lab_data


def read_data():

    file_path = "C:/Users/julia/OneDrive/Documents/FINAL-REPORT-SRLIMS/FINAL-REPORT-SRLIMS/plantilla-reporte-final.xlsx"
    path_file_source = "C:/Users/julia/OneDrive/Documents/FINAL-REPORT-SRLIMS/FINAL-REPORT-SRLIMS/SOURCE-FORMAT.xlsx"
    path_file_write = "C:/Users/julia/OneDrive/Documents/FINAL-REPORT-SRLIMS/FINAL-REPORT-SRLIMS/Reporte.xlsx"

    wb_to_read = load_workbook(filename=file_path, data_only=True)

    wb_to_print= load_workbook(path_file_write)

    wb_to_format = load_workbook(path_file_source)

    # Get the header data

    header_data = excel_header_reader(wb_to_read)

    client_sample_id = header_data[6]

    #First read the chain of custody data

    chain_data = excel_chain_data_reader(wb_to_read, file_path, [23, 23])

    matrix_data = matrix_data_reader(wb_to_read, chain_data)

    last_row = header_format_copy(wb_to_format,wb_to_print,  wb_to_format["Header"], 1)

    write_header_data(wb_to_print,header_data)

    lab_format_copy(wb_to_format, wb_to_print, wb_to_format["Header_lab"], last_row, 20)
    last_row=write_lab_data(wb_to_print, matrix_data, last_row, client_sample_id)
    
    footer_format_copy(wb_to_format, wb_to_print, "Footer", last_row)
    print(f"PUTA ROOW PARA EL FOOTER {last_row}")
    
    last_row = header_format_copy(wb_to_format,wb_to_print,  wb_to_format["Header"], last_row + 7)
    
    header_analitic_format_copy(wb_to_format, wb_to_print, "header_analitic", last_row)
    
    
    set_height_for_all_rows(wb_to_print["Reporte"], 100, 1, None)
    apply_font_to_worksheet(wb_to_print["Reporte"])
    
    wb_to_print.save(path_file_write)

read_data()